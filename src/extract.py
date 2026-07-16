#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fapiao2excel —— 批量把一整个文件夹的票据/发票图片，识别成一张结构化 Excel 表。

和「直接把图丢进千问 App」的区别：App 一次一张、手动传、手动抄；
本工具一条命令跑完整个文件夹，直接落地成带固定字段的 Excel，适合一堆票据要归档的场景。

分层识别（防 AI 幻觉）：
    1) 电子发票 PDF 有文字层 → 直接正则精确解析，0 token、0 幻觉、金额不靠猜
    2) 图片/扫描件（无文字层）→ 才用视觉大模型兜底
输出多一列「识别方式」，一眼看出哪些行是精确解析、哪些是 AI 识别。

用法：
    python src/extract.py <文件夹> [-o 输出.xlsx] [--model 模型名]

依赖：见 requirements.txt。纯电子发票 PDF 无需任何 key；用到 AI 兜底才需 .env 配 key。
"""

import argparse
import base64
import json
import os
import re
import sys
from pathlib import Path

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
PDF_EXTS = {".pdf"}
INPUT_EXTS = IMAGE_EXTS | PDF_EXTS  # 电子发票多为 PDF，一并支持

# 要提取的字段（顺序即 Excel 列顺序）；对准中国大陆常见增值税发票/票据
FIELDS = [
    "发票类型",
    "发票代码",
    "发票号码",
    "开票日期",
    "购买方名称",
    "销售方名称",
    "金额(不含税)",
    "税额",
    "价税合计",
    "备注",
]

PROMPT = (
    "你是财务票据识别助手。请识别这张票据图片，提取下列字段并"
    "严格以 JSON 返回（不要多余解释、不要 markdown 代码块）：\n"
    + "、".join(FIELDS)
    + "。\n规则：金额类字段只保留数字（去掉“¥”和逗号）；"
    "日期统一为 YYYY-MM-DD；某字段图中没有就填空字符串 \"\"；"
    "无法判断发票类型时填“未知”。"
)


def encode_image(path: Path) -> str:
    mime = "image/jpeg"
    ext = path.suffix.lower()
    if ext == ".png":
        mime = "image/png"
    elif ext == ".webp":
        mime = "image/webp"
    elif ext == ".bmp":
        mime = "image/bmp"
    b64 = base64.b64encode(path.read_bytes()).decode("utf-8")
    return f"data:{mime};base64,{b64}"


def pdf_to_data_urls(path: Path, dpi: int = 200) -> list[str]:
    """把 PDF 每一页渲染成 PNG，返回 data_url 列表（电子发票通常 1 页）。"""
    import fitz  # PyMuPDF

    urls = []
    doc = fitz.open(path)
    try:
        for page in doc:
            pix = page.get_pixmap(dpi=dpi)
            b64 = base64.b64encode(pix.tobytes("png")).decode("utf-8")
            urls.append(f"data:image/png;base64,{b64}")
    finally:
        doc.close()
    return urls


def pdf_has_text_layer(path: Path) -> bool:
    import fitz
    doc = fitz.open(path)
    try:
        return any(page.get_text().strip() for page in doc)
    finally:
        doc.close()


# 采信精确解析的硬门槛：发票号码 + 自校验通过的价税合计，两者都有才算数
# （金额没通过「金额+税额=价税合计」自校验时，价税合计为空 → 达不到门槛 → 回退 AI）
_KEY_FIELDS = ("发票号码", "价税合计")


def parse_pdf_textlayer(path: Path) -> tuple[dict | None, int]:
    """从 PDF 文字层直接正则解析发票字段（0 token、0 幻觉）。
    返回 (字段字典 或 None, 命中关键字段数)。无文字层（扫描件）返回 (None, 0)。

    ⚠️ 正则针对「全电发票 / 增值税电子普通发票」常见版式，真实票据版式多样，
    可能需要按你手上的样本微调；命中不足会自动回退到 AI 识别。"""
    import fitz

    doc = fitz.open(path)
    try:
        raw = "\n".join(page.get_text() for page in doc)
        if not raw.strip():
            return None, 0  # 扫描件，无文字层 → 交给 AI
        flat = re.sub(r"[ \t\u00a0]+", " ", raw).replace("\n", " ")
        data = {f: "" for f in FIELDS}
        _fill_textlayer_fields(flat, data)
        # 购买方/销售方靠坐标判定（文字流顺序在不同模板里不可靠），失败再回退文字流顺序
        _fill_parties(doc, flat, data)
    finally:
        doc.close()

    hit = sum(1 for f in _KEY_FIELDS if data[f])
    return data, hit


def _fill_textlayer_fields(flat: str, data: dict) -> None:
    """从扁平化文字层抓 发票号码/代码/日期/金额三元组/发票类型（除买卖方外的字段）。"""

    def grab(pattern, group=1):
        m = re.search(pattern, flat)
        return m.group(group).strip() if m else ""

    # 发票号码：优先「标签后紧跟」（旧版式），否则全电发票 = 独立的 20 位数字
    num = grab(r"发票号码[:：]?\s*(\d{8,20})")
    if not num:
        m = re.search(r"(?<!\d)(\d{20})(?!\d)", flat)
        num = m.group(1) if m else ""
    data["发票号码"] = num
    data["发票代码"] = grab(r"发票代码[:：]?\s*(\d{10,12})")

    # 开票日期：整篇任意位置的「YYYY年MM月DD日」
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", flat)
    if m:
        data["开票日期"] = f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"

    # 金额三元组自校验：金额 + 税额 = 价税合计，对得上才采信（0 幻觉硬保证）
    amts = [float(x.replace(",", "")) for x in re.findall(r"[¥￥]\s*([\d,]+\.\d{2})", flat)]
    triple = _find_amount_triple(amts)
    if triple:
        amount, tax, total = triple
        data["金额(不含税)"] = f"{amount:.2f}"
        data["税额"] = f"{tax:.2f}"
        data["价税合计"] = f"{total:.2f}"

    for kw in ("增值税专用发票", "增值税电子专用发票", "电子发票(普通发票)", "电子发票（普通发票）",
               "增值税电子普通发票", "增值税普通发票", "机动车销售统一发票", "普通发票"):
        if kw in flat:
            data["发票类型"] = kw
            break

    return


def _find_amount_triple(amts: list[float]) -> tuple[float, float, float] | None:
    """从一堆金额里找出满足 金额 + 税额 = 价税合计（金额≥税额）的组合。
    找到即自校验通过 → 金额可精确采信；找不到 → 返回 None 交给 AI。"""
    from itertools import combinations

    if len(amts) < 2:
        return None
    for total in sorted(set(amts), reverse=True):
        smaller = [x for x in amts if x < total - 1e-9]
        for a, b in combinations(smaller, 2):
            if abs(a + b - total) < 0.005:
                return max(a, b), min(a, b), total
    return None


# 公司名后缀（多字后缀，避免像单字"社/行"误命中"统一社会信用代码"）
_COMPANY_TAIL = ("有限公司", "股份公司", "公司", "集团", "中心", "厂", "事务所", "工作室",
                 "个体工商户", "合作社", "银行", "医院", "学院", "大学", "研究院", "事业部")
_COMPANY_RE = r"[\u4e00-\u9fa5（）()]{2,40}?(?:" + "|".join(_COMPANY_TAIL) + r")"
_NAME_NOISE = ("名称", "统一", "识别号", "信息", "项目", "服务费", "价税", "税人")


def _looks_like_company(text: str) -> bool:
    t = (text or "").strip()
    if not (4 <= len(t) <= 40):
        return False
    if any(bad in t for bad in _NAME_NOISE):
        return False
    return t.endswith(_COMPANY_TAIL) and bool(re.fullmatch(r"[\u4e00-\u9fa5（）()]+", t))


def _fill_parties(doc, flat: str, data: dict) -> None:
    """判定 购买方/销售方 名称。
    首选坐标法：用「购」「销」标签坐标 + 各公司名词块坐标就近归属（文字流顺序在不同模板里不可靠，坐标可靠）；
    坐标法拿不到再退回文字流顺序启发式。"""
    buyer = seller = None
    try:
        buyer, seller = _parties_by_coords(doc)
    except Exception:
        buyer = seller = None
    if not (buyer or seller):
        buyer, seller = _parties_by_flow(flat)
    if buyer:
        data["购买方名称"] = buyer
    if seller:
        data["销售方名称"] = seller


def _parties_by_coords(doc):
    """用词块坐标把两个公司名分给 购买方/销售方。返回 (购买方, 销售方) 或 (None, None)。"""
    for page in doc:
        words = page.get_text("words")  # (x0, y0, x1, y1, text, block, line, wno)
        buy_anchor = next(((w[0], w[1]) for w in words if "购" in w[4]), None)
        sell_anchor = next(((w[0], w[1]) for w in words if "销" in w[4]), None)
        if not buy_anchor or not sell_anchor:
            continue
        comp_words = [(w[4].strip(), w[0], w[1]) for w in words if _looks_like_company(w[4])]
        if not comp_words:
            continue
        buyer = seller = None
        best_b = best_s = None
        for name, x, y in comp_words:
            db = abs(x - buy_anchor[0]) + abs(y - buy_anchor[1])
            ds = abs(x - sell_anchor[0]) + abs(y - sell_anchor[1])
            if db <= ds:
                if best_b is None or db < best_b:
                    best_b, buyer = db, name
            elif best_s is None or ds < best_s:
                best_s, seller = ds, name
        if buyer or seller:
            return buyer, seller
    return None, None


def _parties_by_flow(flat: str):
    """回退：按公司名在文字流里出现的先后 + 「购买方/销售方」标签先后尽力分派（不保证准确）。"""
    nospace = flat.replace(" ", "")
    comps = []
    for m in re.finditer(_COMPANY_RE, nospace):
        name = m.group(0).lstrip("日月年号票人注备")  # 去掉可能粘上的日期/标签残字
        if _looks_like_company(name) and name not in comps:
            comps.append(name)
    if not comps:
        return None, None
    ib, isell = nospace.find("购买方"), nospace.find("销售方")
    buyer_first = ib != -1 and (isell == -1 or ib < isell)
    if len(comps) >= 2:
        return (comps[0], comps[1]) if buyer_first else (comps[1], comps[0])
    return comps[0], None


def parse_json_loose(text: str) -> dict:
    """模型偶尔会包 ```json 代码块或前后带字，尽量抠出 JSON。"""
    t = text.strip()
    if t.startswith("```"):
        t = t.strip("`")
        if t[:4].lower() == "json":
            t = t[4:]
    start, end = t.find("{"), t.rfind("}")
    if start != -1 and end != -1 and end > start:
        t = t[start : end + 1]
    return json.loads(t)


def extract_one(client, model: str, data_url: str) -> dict:
    resp = client.chat.completions.create(
        model=model,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": PROMPT},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            }
        ],
        temperature=0,
    )
    content = resp.choices[0].message.content or ""
    try:
        data = parse_json_loose(content)
    except Exception:
        # 解析失败：整条原文塞进备注，保证这一行不丢
        data = {f: "" for f in FIELDS}
        data["备注"] = f"[解析失败] {content[:200]}"
    return data


def write_excel(rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "票据明细"

    headers = ["文件名"] + FIELDS + ["识别方式"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D6F4E6")
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append(
            [r.get("文件名", "")]
            + [str(r.get(f, "")) for f in FIELDS]
            + [r.get("识别方式", "")]
        )

    # 简单自适应列宽
    for col in range(1, len(headers) + 1):
        letter = ws.cell(row=1, column=col).column_letter
        longest = max(
            (len(str(ws.cell(row=i, column=col).value or "")) for i in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 40)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="批量把票据/发票图片识别成一张结构化 Excel 表（引擎：通义千问 VL，自带 key）"
    )
    ap.add_argument("folder", help="存放票据图片的文件夹")
    ap.add_argument("-o", "--out", default="out/票据明细.xlsx", help="输出 Excel 路径")
    ap.add_argument("--model", default=None, help="覆盖 .env 里的 QWEN_VL_MODEL")
    args = ap.parse_args()

    folder = Path(args.folder)
    if not folder.is_dir():
        print(f"✗ 文件夹不存在：{folder}", file=sys.stderr)
        return 1

    files = sorted(p for p in folder.iterdir() if p.suffix.lower() in INPUT_EXTS)
    if not files:
        print(f"✗ {folder} 里没找到票据文件（支持 {', '.join(sorted(INPUT_EXTS))}）", file=sys.stderr)
        return 1

    try:
        from dotenv import load_dotenv
        load_dotenv()
    except Exception:
        pass

    # AI 客户端「按需」创建：能精确解析的 PDF 完全不碰 AI，也就不需要 key
    _client = {"obj": None}
    base_url = os.getenv("BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1")
    model = args.model or os.getenv("VL_MODEL") or os.getenv("QWEN_VL_MODEL", "qwen-vl-max")

    def get_client():
        if _client["obj"] is None:
            api_key = os.getenv("API_KEY") or os.getenv("DASHSCOPE_API_KEY")
            if not api_key:
                raise RuntimeError(
                    "需要用 AI 识别（图片/扫描件），但没读到 API_KEY。"
                    "把 .env.example 复制成 .env 填入 key，或只放电子发票 PDF（0 token 不需要 key）。"
                )
            from openai import OpenAI
            _client["obj"] = OpenAI(api_key=api_key, base_url=base_url)
        return _client["obj"]

    rows = []
    n_exact = n_ai = n_fail = 0
    total = len(files)
    print(f"共 {total} 个文件，开始识别（PDF 优先精确解析，图片/扫描件用 AI 兜底）…")

    for i, f in enumerate(files, start=1):
        print(f"  [{i}/{total}] {f.name} …", end=" ", flush=True)
        try:
            ext = f.suffix.lower()
            # ── 第一层：PDF 文字层精确解析（0 token、0 幻觉）──
            if ext in PDF_EXTS:
                parsed, hit = parse_pdf_textlayer(f)
                if parsed is not None and hit >= 2:
                    parsed["文件名"] = f.name
                    parsed["识别方式"] = "精确解析(PDF)"
                    rows.append(parsed)
                    n_exact += 1
                    print("✓ 精确解析")
                    continue
                # 文字层缺失/不足 → 渲染成图交给 AI
                data_urls = pdf_to_data_urls(f)
                pages = ([(f.name, data_urls[0])] if len(data_urls) == 1
                         else [(f"{f.name} (第{j}页)", u) for j, u in enumerate(data_urls, 1)])
            # ── 图片 ──
            else:
                pages = [(f.name, encode_image(f))]

            # ── 第二层：AI 视觉识别 ──
            for name, data_url in pages:
                data = extract_one(get_client(), model, data_url)
                data["文件名"] = name
                data["识别方式"] = "AI识别"
                rows.append(data)
                n_ai += 1
            print("✓ AI")
        except Exception as e:
            print(f"✗ {e}")
            rows.append({"文件名": f.name, "备注": f"[识别失败] {e}", "识别方式": "失败"})
            n_fail += 1

    out_path = Path(args.out)
    write_excel(rows, out_path)
    print(f"\n完成 → {out_path.resolve()}")
    print(f"统计：精确解析 {n_exact} · AI识别 {n_ai} · 失败 {n_fail}")
    if n_ai:
        print(f"（AI 识别用的模型：{model}；金额税号等重要字段建议人工复核）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
